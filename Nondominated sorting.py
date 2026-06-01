import pandas as pd
import numpy as np
import os
import gc
import time
import pyarrow as pa
import pyarrow.parquet as pq
import openpyxl
import sys
import logging
from scipy.spatial import cKDTree
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import seaborn as sns

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()


def setup_logging():
    """设置详细日志记录"""
    logger.setLevel(logging.INFO)
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    return logger


def efficient_pareto_frontier(points):
    """高效三维帕累托前沿计算，正确识别非支配解"""
    if len(points) == 0:
        return np.array([], dtype=bool)

    # 创建点索引数组
    indices = np.arange(len(points))
    is_efficient = np.ones(len(points), dtype=bool)

    # 按第一个目标排序（升序）
    sort_idx = np.argsort(points[:, 0])
    points = points[sort_idx]
    indices = indices[sort_idx]

    # 使用KDTree加速邻域搜索
    tree = cKDTree(points)

    # 检查每个点是否被支配
    for i in range(len(points)):
        if not is_efficient[i]:
            continue

        # 查找可能支配当前点的候选点（在第一个目标上更优的点）
        candidate_indices = tree.query_ball_point(points[i], r=np.inf, p=np.inf)

        for j in candidate_indices:
            if j == i or not is_efficient[j]:
                continue

            # 检查点j是否支配点i
            if np.all(points[j] <= points[i]) and np.any(points[j] < points[i]):
                is_efficient[i] = False
                break

    # 恢复原始顺序
    result = np.zeros(len(points), dtype=bool)
    result[sort_idx] = is_efficient

    # 统计结果
    total_dominated = len(points) - np.sum(result)
    logging.info(f"检测到支配点: {total_dominated}/{len(points)} ({total_dominated / len(points):.1%})")

    return result


def read_all_data(input_files, chunk_size=5000):
    """读取所有输入文件的数据并合并"""
    all_data = []

    for file_idx, file_path in enumerate(input_files):
        if not os.path.exists(file_path):
            logger.warning(f"文件不存在，跳过: {file_path}")
            continue

        logger.info(f"[{file_idx + 1}/{len(input_files)}] 读取文件: {os.path.basename(file_path)}")

        # 使用openpyxl进行高效读取
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet = wb.active

            # 获取列名
            col_names = []
            for idx, row in enumerate(sheet.iter_rows(max_row=1, values_only=True)):
                if idx == 0:
                    col_names = [str(cell) for cell in row]
                    break

            rows = []
            processed = 0

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                rows.append(row)
                processed += 1

                # 分批处理
                if len(rows) >= chunk_size:
                    df_chunk = pd.DataFrame(rows, columns=col_names)

                    # 添加数据源标记
                    source_name = os.path.splitext(os.path.basename(file_path))[0]
                    df_chunk['Data_Source'] = source_name

                    all_data.append(df_chunk)
                    rows = []
                    gc.collect()

            # 处理剩余行
            if rows:
                df_chunk = pd.DataFrame(rows, columns=col_names)
                source_name = os.path.splitext(os.path.basename(file_path))[0]
                df_chunk['Data_Source'] = source_name
                all_data.append(df_chunk)

            wb.close()
            logger.info(f"文件处理完成: {processed}行")

        except Exception as e:
            logger.error(f"读取文件错误: {file_path} - {str(e)}")

    # 合并所有数据
    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        logger.info(f"所有文件合并完成: 总行数={len(combined)}")

        # 确保Cluster ID列为字符串
        if 'Cluster ID' in combined.columns:
            combined['Cluster ID'] = combined['Cluster ID'].astype(str)

        return combined
    else:
        logger.error("没有可用的数据")
        return pd.DataFrame()


def compute_global_pareto(data, objectives):
    """在整个数据集上计算全局帕累托前沿"""
    if data.empty:
        logger.error("数据集为空，无法计算帕累托前沿")
        return data

    logger.info(f"开始全局帕累托计算 - 总样本数: {len(data):,}")
    start_time = time.time()

    # 提取目标值
    points = data[objectives].values

    # UDI转换为最小化问题
    if 'UDI' in objectives:
        udi_idx = objectives.index('UDI')
        points[:, udi_idx] = -points[:, udi_idx]

    # 计算帕累托解
    is_pareto = efficient_pareto_frontier(points)

    # 添加标记
    data['Is_Global_Pareto'] = is_pareto

    pareto_count = is_pareto.sum()
    elapsed = time.time() - start_time
    logger.info(f"全局帕累托计算完成: 解数量={pareto_count}/{len(data)} | "
                f"用时={elapsed:.2f}秒 | {(elapsed / len(data)) * 1e6:.2f}微秒/行")

    return data


def save_global_pareto_to_excel(global_pareto_df, output_file="Global_Pareto_Front.xlsx", max_rows=100000):
    """将全局帕累托解保存到Excel文件，处理大数据量情况"""
    if global_pareto_df.empty:
        logger.warning("没有全局帕累托解可供保存")
        return False

    try:
        # 如果数据量过大，分割为多个工作表
        num_rows = len(global_pareto_df)
        logger.info(f"准备保存全局帕累托解到Excel: {num_rows}行")

        if num_rows <= max_rows:
            # 直接保存单个Excel文件
            global_pareto_df.to_excel(output_file, index=False)
            logger.info(f"保存到单文件Excel: {output_file}")
            return True

        # 大型数据集处理：创建多个工作表
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        num_sheets = (num_rows + max_rows - 1) // max_rows
        logger.info(f"数据量过大({num_rows}行)，创建{num_sheets}个工作表")

        for i in range(num_sheets):
            start_idx = i * max_rows
            end_idx = min((i + 1) * max_rows, num_rows)
            sheet_name = f"ParetoSheet_{i + 1}"

            # 创建当前工作表的子集
            df_sub = global_pareto_df.iloc[start_idx:end_idx].copy()
            df_sub.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(f"保存工作表 {sheet_name}: {len(df_sub)}行")

        writer.close()
        logger.info(f"保存完成: {output_file}")
        return True

    except Exception as e:
        logger.error(f"保存Excel失败: {str(e)}")
        return False


def visualize_pareto_front(global_pareto_df, objectives, output_dir="pareto_visualizations"):
    """
    Create comprehensive visualizations of the Pareto front
    Including 3D view, scatter matrix, histograms, and heatmaps
    Note: UDI values are restored to their original scale (positive) for visualization
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Create a copy to avoid modifying original data
    plot_df = global_pareto_df.copy()

    # Restore UDI to original positive values for visualization
    if 'UDI' in plot_df:
        plot_df['UDI'] = -plot_df['UDI']

    # Create color palette for Data_Source
    sources = plot_df['Data_Source'].unique()
    num_sources = len(sources)
    palette = sns.color_palette("husl", num_sources)
    color_dict = dict(zip(sources, palette))

    # 1. 3D Pareto Front Visualization
    plt.figure(figsize=(12, 10))
    ax = plt.axes(projection='3d')

    # Extract objectives
    x = plot_df[objectives[0]]
    y = plot_df[objectives[1]]
    z = plot_df['UDI'] if 'UDI' in objectives else plot_df[objectives[2]]  # 直接使用UDI列

    # Plot each data source with different color
    for source in sources:
        source_data = plot_df[plot_df['Data_Source'] == source]
        ax.scatter3D(
            source_data[objectives[0]],
            source_data[objectives[1]],
            source_data[objectives[2]],
            color=color_dict[source],
            label=source,
            s=25,
            alpha=0.7
        )

    ax.set_xlabel(objectives[0], fontsize=12, labelpad=10)
    ax.set_ylabel(objectives[1], fontsize=12, labelpad=10)
    ax.set_zlabel(objectives[2], fontsize=12, labelpad=10)
    ax.set_title('3D Pareto Front', fontsize=14, pad=20)

    # Add legend outside the plot
    ax.legend(loc='upper left', bbox_to_anchor=(0.7, 0.9), fontsize=9)

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "3D_pareto_front.png"), dpi=300)
    plt.close()
    logger.info("Created 3D Pareto front visualization")

    # 2. Pairwise Scatter Matrix with Data Source Coloring
    plt.figure(figsize=(15, 12))
    sns.set(style="ticks", palette="husl", font_scale=1.0)

    # Create scatter matrix
    g = sns.pairplot(
        plot_df,
        vars=objectives,
        hue='Data_Source',
        palette=color_dict,
        diag_kind="kde",
        plot_kws=dict(alpha=0.5, s=15),
        diag_kws=dict(fill=True, alpha=0.3),
        corner=False
    )

    g.fig.suptitle('Pareto Solutions - Objective Relationships', fontsize=16, y=1.02)

    # Move legend outside
    g._legend.set_bbox_to_anchor((1.0, 0.5))

    plt.savefig(os.path.join(output_dir, "pairwise_scatter_matrix.png"), dpi=300, bbox_inches='tight')
    plt.close()
    logger.info("Created pairwise scatter matrix")

    # 3. Histograms of Objectives
    plt.figure(figsize=(16, 5))
    for i, obj in enumerate(objectives):
        plt.subplot(1, 3, i + 1)

        # Create histogram with different colors for each data source
        for source in sources:
            source_data = plot_df[plot_df['Data_Source'] == source]
            sns.histplot(
                source_data[obj],
                kde=True,
                color=color_dict[source],
                label=source,
                alpha=0.4,
                bins=30,
                element="step",
                stat="density"
            )

        plt.title(f"{obj} Distribution", fontsize=12)
        plt.xlabel(obj, fontsize=10)
        plt.ylabel("Density", fontsize=10)
        plt.grid(alpha=0.3)

        # Add legend to the first subplot
        if i == 0:
            plt.legend(fontsize=8, title='Data Source', title_fontsize=9)

    plt.suptitle('Objective Distributions of Pareto Solutions', fontsize=16, y=1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "objective_histograms.png"), dpi=300)
    plt.close()
    logger.info("Created objective histograms")

    # 4. Correlation Heatmap
    plt.figure(figsize=(10, 8))
    sns.heatmap(
        plot_df[objectives].corr(),
        annot=True,
        cmap='coolwarm',
        fmt=".2f",
        vmin=-1,
        vmax=1,
        linewidths=0.5,
        annot_kws={"size": 12}
    )

    plt.title('Objective Correlation Heatmap', fontsize=16)
    plt.savefig(os.path.join(output_dir, "objective_correlation.png"), dpi=300)
    plt.close()
    logger.info("Created objective correlation heatmap")

    # 5. Solution Source Distribution
    plt.figure(figsize=(8, 6))
    source_counts = plot_df['Data_Source'].value_counts()
    plt.pie(
        source_counts,
        labels=source_counts.index,
        autopct='%1.1f%%',
        colors=palette,
        startangle=90,
        wedgeprops=dict(width=0.6, edgecolor='w')
    )
    plt.title('Distribution of Pareto Solutions by Data Source', fontsize=14)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "solution_source_distribution.png"), dpi=300)
    plt.close()
    logger.info("Created solution source distribution pie chart")

    logger.info(f"All visualizations saved to: {os.path.abspath(output_dir)}")


def visualize_cluster_distribution(final_data, output_dir="pareto_visualizations"):
    """Visualize distribution of Pareto solutions across clusters"""
    if 'Cluster ID' not in final_data or 'Is_Global_Pareto' not in final_data:
        return

    # Prepare data for Pareto solutions only
    pareto_data = final_data[final_data['Is_Global_Pareto']].copy()

    # Create color palette for clusters
    clusters = pareto_data['Cluster ID'].unique()
    num_clusters = len(clusters)
    cluster_palette = sns.color_palette("Set2", num_clusters)

    plt.figure(figsize=(14, 6))

    # Plot 1: Count of Pareto solutions per cluster
    plt.subplot(1, 2, 1)
    cluster_counts = pareto_data['Cluster ID'].value_counts().sort_index()
    bars = plt.bar(
        cluster_counts.index.astype(str),
        cluster_counts.values,
        color=cluster_palette
    )

    plt.title('Pareto Solutions per Cluster', fontsize=14)
    plt.xlabel('Cluster ID', fontsize=12)
    plt.ylabel('Number of Pareto Solutions', fontsize=12)
    plt.xticks(rotation=45)
    plt.grid(axis='y', alpha=0.3)

    # Add labels
    for bar in bars:
        height = bar.get_height()
        plt.annotate(f'{height}',
                     xy=(bar.get_x() + bar.get_width() / 2, height),
                     xytext=(0, 3),  # 3 points vertical offset
                     textcoords="offset points",
                     ha='center', va='bottom', fontsize=9)

    # Plot 2: Pareto ratio per cluster
    plt.subplot(1, 2, 2)
    cluster_ratio = (
        final_data.groupby('Cluster ID')['Is_Global_Pareto']
        .mean()
        .sort_index()
    )

    bars = plt.bar(
        cluster_ratio.index.astype(str),
        cluster_ratio.values,
        color=cluster_palette
    )

    plt.title('Pareto Solution Ratio per Cluster', fontsize=14)
    plt.xlabel('Cluster ID', fontsize=12)
    plt.ylabel('Percentage of Pareto Solutions', fontsize=12)
    plt.xticks(rotation=45)
    plt.grid(axis='y', alpha=0.3)

    # Add percentage labels
    for bar in bars:
        height = bar.get_height()
        plt.annotate(f'{height:.1%}',
                     xy=(bar.get_x() + bar.get_width() / 2, height),
                     xytext=(0, 3),
                     textcoords="offset points",
                     ha='center', va='bottom', fontsize=9)

    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "cluster_distribution_guided set.png"), dpi=300)
    plt.close()
    logger.info("Created cluster distribution visualization")


def main():
    """主处理函数 - 针对大型数据集优化"""
    logger = setup_logging()
    logger.info("开始高效全局帕累托前沿计算...")

    # 配置
    input_files = [
        'Cluster 0_New.xlsx',
        'Cluster 1_New.xlsx',
        'Cluster 2_New.xlsx',
         'Cluster 3_New.xlsx',
         'Cluster 9_New.xlsx',
         'Cluster 12_New.xlsx',
         'Cluster 15_New.xlsx',
         'Cluster 24_New.xlsx'
    ]
    output_file = 'Global_Pareto_Front_guided set.parquet'
    objectives = ['EUI', 'PPD', 'UDI']

    # 阶段1: 读取并合并所有文件
    logger.info("阶段1: 读取并合并所有输入文件...")
    start_time = time.time()

    # 读取所有数据
    combined_data = read_all_data(input_files)
    if combined_data.empty:
        logger.error("没有读取到有效数据，程序终止")
        return

    # 报告内存使用情况
    mem_usage = combined_data.memory_usage(deep=True).sum() / (1024 ** 2)  # MB
    logger.info(f"合并数据集内存使用量: {mem_usage:.2f} MB | 总行数: {len(combined_data):,}")

    # 阶段2: 在整个数据集上计算全局帕累托前沿
    logger.info("\n阶段2: 计算全局帕累托前沿...")

    # 计算全局帕累托前沿
    combined_data = compute_global_pareto(combined_data, objectives)

    # 提取全局帕累托解
    global_pareto = combined_data[combined_data['Is_Global_Pareto']].copy()
    logger.info(f"全局帕累托解数量: {len(global_pareto):,} (占比 {len(global_pareto) / len(combined_data):.2%})")

    # 阶段3: 保存结果
    logger.info("\n阶段3: 保存结果...")

    # 保存到Parquet
    table = pa.Table.from_pandas(combined_data, preserve_index=False)
    pq.write_table(table, output_file, compression='ZSTD')
    file_size = os.path.getsize(output_file) / (1024 ** 2)
    logger.info(f"保存完整数据集到Parquet: {output_file} | 文件大小: {file_size:.2f} MB")

    # 保存全局帕累托解到Excel（恢复UDI原始值）
    global_pareto_excel = global_pareto.copy()
    if 'UDI' in objectives:
        global_pareto_excel['UDI'] = -global_pareto_excel['UDI']  # 恢复UDI原始正值

    save_global_pareto_to_excel(
        global_pareto_excel,
        output_file="Global_Pareto_Solutions_guided set.xlsx"
    )

    # 阶段4: 可视化
    logger.info("\n阶段4: 创建可视化...")
    if not global_pareto_excel.empty:
        visualize_pareto_front(global_pareto_excel, objectives)
        visualize_cluster_distribution(combined_data)

    # 阶段5: 最终报告
    elapsed_total = time.time() - start_time
    logger.info(f"\n{'=' * 50}")
    logger.info(f"处理完成! 总用时: {elapsed_total:.2f}秒")
    logger.info(f"总样本数: {len(combined_data):,}")
    logger.info(f"全局帕累托解: {len(global_pareto):,}")
    logger.info(f"结果保存位置:")
    logger.info(f"  - 完整数据集: {output_file}")
    logger.info(f"  - 帕累托解: Global_Pareto_Solutions.xlsx")
    logger.info(f"  - 可视化: pareto_visualizations/ 目录")
    logger.info("=" * 50)


if __name__ == "__main__":
    logger = setup_logging()

    # 检查依赖
    try:
        import pandas as pd
        import numpy as np
        import openpyxl
        import pyarrow
        from scipy.spatial import cKDTree
    except ImportError as e:
        logger.error(f"缺少依赖库: {str(e)}")
        logger.info("请运行: pip install pandas numpy openpyxl pyarrow scipy")
        sys.exit(1)

    try:
        main()
    except MemoryError:
        logger.error("内存不足! 尝试减小chunk_size或增加系统内存")
        sys.exit(1)
    except Exception as e:
        logger.exception(f"处理过程中出错: {str(e)}")
        sys.exit(1)